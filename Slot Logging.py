import pyautogui
from PIL import Image, ImageOps, ImageEnhance
import pytesseract
import openpyxl
from datetime import datetime
import time

# Tesseract exe locaiton
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# Excel File named
excel_file = "Slot_result.xlsx"

# Image area numbers
total_credit_area = (176, 966, 479, 1004)
total_win_area = (695, 966, 800, 1004)
total_bet_area = (1164, 966, 1227, 1004)

# Auto click coordinates
spin_button = (1611, 978)
continue_button = (957, 879)
up_wager = (1365, 984)
down_wager = (1023, 984)

# Preprocessing 
def preprocessing(img):
    img = ImageOps.grayscale(img)
    img = ImageEnhance.Contrast(img).enhance(2.0)
    img = img.point(lambda x:0 if x < 128 else 255, '1')
    return img


# Helper function
def extract_the_numbers (img):
    """
    Extract integer number from image using OCR
    """
    text = pytesseract.image_to_string(
        img,
        config = '--psm 7 -c tessedit_char_whitelist=0123456789.,'
    ).strip()
    cleaned_text = ''.join(filter(str.isdigit,text))
    if not cleaned_text:
        return 0
    return int(cleaned_text)

# Capture images
def capture_credit():
    screenshot = pyautogui.screenshot()
    credit_img = preprocessing(screenshot.crop(total_credit_area))
    return extract_the_numbers(credit_img)

def capture_bet_win():
    screenshot = pyautogui.screenshot()

    bet_img = preprocessing(screenshot.crop(total_bet_area))
    win_img = preprocessing(screenshot.crop(total_win_area))

    bet = extract_the_numbers(bet_img)
    win = extract_the_numbers(win_img)

    return bet, win

# Delay for balance changes for animation
def stable_balance (capture_func, check_delay =0.3, stable_count = 3):
    last_value = capture_func()
    stable_time = 0

    while stable_time < stable_count:
        time.sleep(check_delay)
        new_value = capture_func()

        if new_value == last_value:
            stable_time += 1
        else:
            stable_time = 0
            last_value = new_value
    return last_value


# Excel format
try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Game No.", "Gamemode" , "Prev Balance", "Bet", "Total Win", "Expected Balance",
            "Actual Balance", "Difference", "Status", "Timestamp"])
    wb.save(excel_file)
    ws = wb.active


# game number (auto increment) function
def next_game ():
    if ws.max_row < 2 : 
        return 1
    return ws.cell(row=ws.max_row, column = 1).value + 1
 
# Spin counter
spin_counter = -1
action_delay = 2.5 # Delay timer between each auto clicker action
bet_step = 0 # Counter for the clicker to go for bet adjustment


# Main Loop
print ("Starting the slot logging capture.... Press Ctrl+C to stop")
prev_balance = None
last_logged_balance = None
tolerance = 2

current_mode = "Normal"
total_run = int(input("Enter number of spins you want : "))
while spin_counter < total_run:

    spin_counter += 1

    # Adjust the bet 
    if spin_counter % 30 == 1:  # every 30 spins, first spin triggers adjustment
        if bet_step == 0 or bet_step == 1:
            pyautogui.moveTo(up_wager)
            pyautogui.click()
            print(f"Spin {spin_counter}: Up Bet")
        elif bet_step == 2 or bet_step == 3:
            pyautogui.moveTo(down_wager)
            pyautogui.click()
            print(f"Spin {spin_counter}: Down Bet")
        time.sleep(action_delay)

    balance_before = capture_credit()

    pyautogui.moveTo(spin_button)
    pyautogui.click()
    print(f"Spin {spin_counter}: Spinz za wheel")
    time.sleep(3)

    balance_after = capture_credit()
    if balance_after < balance_before:
        current_mode = "Normal"
    else:
        current_mode = "Jackpot"

    print(f"Mode : {current_mode}")

    time.sleep(action_delay)

    pyautogui.moveTo(continue_button)
    pyautogui.click()
    print(f"Spin {spin_counter}: Continued")
    time.sleep(action_delay)

    if spin_counter % 30 == 0:
        bet_step = (bet_step + 1) % 4

    current_credit = capture_credit()

    if prev_balance is None:
        prev_balance = current_credit
        last_logged_balance = current_credit
        print (f"Previous Balance : {prev_balance}")
        time.sleep(1)
        continue
    
    if current_credit != last_logged_balance and current_credit != 0:
        final_credit = stable_balance(capture_credit)
        
        if final_credit == last_logged_balance:
            continue

        bet, win = capture_bet_win()

         # Game number
        game = next_game()
    
        # Gamemode based
        if current_mode == "Normal":
            expected = prev_balance - bet + win
        else:
            expected = prev_balance + win

        # Difference 
        diff = expected - final_credit

        # Status
        if  abs(diff) <= tolerance:
            status ="OK"
            diff = 0
        else:
            status = "Incorrect"

        # Convert None to 0 / 0 is mean value is empty
        bet = bet if bet is not None else 0
        win = win if win is not None else 0
        final_credit = final_credit if final_credit is not None else 0

        # Append to Excel
        ws.append([
            game,
            current_mode,
            prev_balance,
            bet,
            win,
            expected,
            final_credit,
            diff,
            status,
            datetime.now()
        ])
        wb.save(excel_file)

        print(f"Game {game} | Gamemode: {current_mode}: | Prev: {prev_balance} | Bet: {bet} | Win: {win} | "
            f"Expected Balance : {expected} | Actual Balance : {final_credit} | Difference : {diff} | Status: {status}")
        
        prev_balance = final_credit
        last_logged_balance = final_credit
    print ("Spinning is completed.")
    time.sleep(1)
